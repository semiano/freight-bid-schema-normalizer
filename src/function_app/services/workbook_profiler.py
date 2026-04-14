from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..models.contracts import SheetProfile, WorkbookProfile
from .sheet_classifier import classify_sheet


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _infer_type(values: list[Any]) -> str:
    non_null_values = [value for value in values if value not in (None, "")]
    if not non_null_values:
        return "unknown"

    if all(isinstance(value, bool) for value in non_null_values):
        return "bool"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in non_null_values):
        return "int"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in non_null_values):
        return "float"

    return "str"


def _header_score(row_values: list[Any]) -> int:
    normalized = [_stringify(value) for value in row_values]
    non_empty = [value for value in normalized if value]
    if not non_empty:
        return -1

    unique_count = len(set(non_empty))
    alpha_count = len([value for value in non_empty if any(character.isalpha() for character in value)])
    return (len(non_empty) * 2) + unique_count + alpha_count


def _detect_header_row_candidates(sheet: Worksheet, scan_limit: int = 30) -> list[int]:
    candidates: list[tuple[int, int]] = []
    max_scan_row = min(sheet.max_row or 1, scan_limit)

    for row_index in range(1, max_scan_row + 1):
        row_values = [sheet.cell(row=row_index, column=column_index).value for column_index in range(1, (sheet.max_column or 1) + 1)]
        score = _header_score(row_values)
        if score > 0:
            candidates.append((row_index, score))

    candidates.sort(key=lambda item: item[1], reverse=True)
    return [row_index for row_index, _ in candidates[:3]]


def _extract_columns(sheet: Worksheet, header_row: int) -> list[str]:
    columns: list[str] = []
    for column_index in range(1, (sheet.max_column or 1) + 1):
        header_value = _stringify(sheet.cell(row=header_row, column=column_index).value)
        if header_value:
            columns.append(header_value)
    return columns


def _extract_sample_rows(sheet: Worksheet, header_row: int, columns: list[str], sample_size: int = 10) -> list[dict[str, Any]]:
    sample_rows: list[dict[str, Any]] = []
    if not columns:
        return sample_rows

    for row_index in range(header_row + 1, (sheet.max_row or header_row) + 1):
        row_data: dict[str, Any] = {}
        has_value = False

        for column_index, column_name in enumerate(columns, start=1):
            value = sheet.cell(row=row_index, column=column_index).value
            row_data[column_name] = value
            if value not in (None, ""):
                has_value = True

        if has_value:
            sample_rows.append(row_data)
        if len(sample_rows) >= sample_size:
            break

    return sample_rows


def _find_duplicate_headers(columns: list[str]) -> list[str]:
    counts = Counter([column.strip().lower() for column in columns if column.strip()])
    duplicates = [name for name, count in counts.items() if count > 1]
    return sorted(duplicates)


def _calculate_empty_column_ratio(columns: list[str], sample_rows: list[dict[str, Any]]) -> float:
    if not columns:
        return 1.0
    if not sample_rows:
        return 1.0

    empty_columns = 0
    for column in columns:
        if all(row.get(column) in (None, "") for row in sample_rows):
            empty_columns += 1

    return empty_columns / len(columns)


def profile_workbook(workbook_path: str, sample_size: int = 10) -> WorkbookProfile:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet_profiles: list[SheetProfile] = []

        for sheet in workbook.worksheets:
            header_candidates = _detect_header_row_candidates(sheet)
            selected_header = header_candidates[0] if header_candidates else 1

            columns = _extract_columns(sheet, selected_header)
            sample_rows = _extract_sample_rows(sheet, selected_header, columns, sample_size=sample_size)

            inferred_types: dict[str, str] = {}
            for column in columns:
                values = [row.get(column) for row in sample_rows]
                inferred_types[column] = _infer_type(values)

            classification = classify_sheet(sheet.title, columns, sample_rows)
            duplicate_headers = _find_duplicate_headers(columns)
            empty_ratio = _calculate_empty_column_ratio(columns, sample_rows)

            notes = f"classification_score={classification['score']}"
            profile = SheetProfile(
                name=sheet.title,
                visible=(sheet.sheet_state == "visible"),
                used_range=sheet.calculate_dimension(),
                header_row=selected_header,
                header_row_candidates=header_candidates,
                columns=columns,
                inferred_types=inferred_types,
                sample_rows=sample_rows,
                duplicate_headers=duplicate_headers,
                empty_column_ratio=empty_ratio,
                likely_business_meaning=classification["business_meaning"],
                classifier_hints=classification["hints"],
                notes=notes,
            )
            sheet_profiles.append(profile)

        workbook_name = Path(workbook_path).name
        return WorkbookProfile(
            workbook_name=workbook_name,
            sheets=sheet_profiles,
            notes=f"profiled_sheet_count={len(sheet_profiles)}",
        )
    finally:
        workbook.close()
