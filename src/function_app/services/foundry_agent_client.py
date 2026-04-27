from __future__ import annotations

import json
import os
import time
import urllib.error
import urllib.request
from typing import Any

from azure.core.credentials import AzureKeyCredential
from azure.identity import DefaultAzureCredential


class FoundryAgentClient:
    def __init__(
        self,
        endpoint: str | None = None,
        agent_name: str | None = None,
        agent_version: str | None = None,
        api_key: str | None = None,
        mode: str = "mock",
    ) -> None:
        self.endpoint = endpoint or os.getenv("FOUNDRY_PROJECT_ENDPOINT", "")
        self.agent_name = agent_name or os.getenv("FOUNDRY_AGENT_NAME", "")
        self.agent_version = (agent_version or os.getenv("FOUNDRY_AGENT_VERSION", "")).strip()
        self.assistant_id = os.getenv("FOUNDRY_ASSISTANT_ID", "")
        self.api_key = api_key or os.getenv("FOUNDRY_API_KEY", "")
        self.mode = mode
        self.last_invocation_report: dict[str, Any] = {
            "mode": mode,
            "path": "uninitialized",
            "agent_name": self.agent_name,
            "agent_version": self.agent_version,
            "assistant_id": self.assistant_id,
        }

    def get_last_invocation_report(self) -> dict[str, Any]:
        return dict(self.last_invocation_report)

    def plan(self, system_prompt: str, user_prompt: str) -> dict[str, Any]:
        if self.mode == "mock":
            self.last_invocation_report = {
                "mode": "mock",
                "path": "mock_response",
                "agent_name": self.agent_name,
                "agent_version": self.agent_version,
                "assistant_id": self.assistant_id,
            }
            return self._mock_response()
        return self._live_response(system_prompt, user_prompt)

    def _mock_response(self) -> dict[str, Any]:
        return {
            "relevant_sheets": ["SINGLE FTL LIVE LOAD TRAILERS"],
            "ignored_sheets": ["REQUIREMENTS", "FSC", "LOCATIONS"],
            "mapping_plan": {
                "Customer Lane ID": "derived",
                "FO Code": "constant:RXOCode",
                "Origin City": "input:ORIGIN CITY",
                "Destination City": "input:DESTINATION CITY",
                "Origin Note": "input:ORIGIN NOTE|ORIGIN NOTES|ORIGIN COMMENT|ORIGIN REMARKS",
                "Destination Note": "input:DESTINATION NOTE|DESTINATION NOTES|DESTINATION COMMENT|DESTINATION REMARKS",
                "Bid Note": "input:BID NOTE|BID NOTES|CARRIER NOTES|NOTES|COMMENTS|REMARKS|LANE NOTES",
            },
            "constants": {
                "FO Code": "RXOCode",
                "FSC Type": "PerMileAmount",
            },
            "enrichments": {
                "Origin Country": "normalize_country",
                "Destination Country": "normalize_country",
            },
            "assumptions": ["Only primary live load sheet is included for v1 planner mock."],
            "confidence_scores": {"overall": 0.72},
            "python_script": """import json
from openpyxl import load_workbook

def _norm_header(value):
    if value is None:
        return ""
    return str(value).strip().upper()

NOTE_FIELDS = [
    ("Origin Note", [
        "ORIGIN NOTE", "ORIGIN NOTES", "ORIGIN COMMENT", "ORIGIN COMMENTS",
        "ORIGIN REMARK", "ORIGIN REMARKS", "ORIGIN INSTRUCTION", "ORIGIN INSTRUCTIONS",
    ]),
    ("Destination Note", [
        "DESTINATION NOTE", "DESTINATION NOTES", "DEST NOTE", "DEST NOTES",
        "DESTINATION COMMENT", "DESTINATION COMMENTS", "DESTINATION REMARK",
        "DESTINATION REMARKS", "DEST INSTRUCTION", "DEST INSTRUCTIONS",
    ]),
    ("Bid Note", [
        "BID NOTE", "BID NOTES", "CARRIER NOTE", "CARRIER NOTES",
        "LANE NOTE", "LANE NOTES", "NOTES", "NOTE", "COMMENTS", "COMMENT",
        "REMARKS", "REMARK",
    ]),
]

def _row_to_record(row, index_map, row_number):
    def val(*names):
        for name in names:
            if name in index_map:
                value = row[index_map[name]]
                if value is None:
                    return ""
                return value
        return ""

    origin_country = val("ORIGIN COUNTRY", "ORIGIN_COUNTRY", "ORIGIN CNTRY") or "US"
    dest_country = val("DESTINATION COUNTRY", "DESTINATION_COUNTRY", "DEST COUNTRY", "DEST CNTRY") or "US"

    origin_note = val(*NOTE_FIELDS[0][1])
    dest_note = val(*NOTE_FIELDS[1][1])
    bid_note = val(*NOTE_FIELDS[2][1])

    notes_array = []
    if origin_note and str(origin_note).strip():
        notes_array.append({"field": "Origin Note", "value": str(origin_note).strip()})
    if dest_note and str(dest_note).strip():
        notes_array.append({"field": "Destination Note", "value": str(dest_note).strip()})
    if bid_note and str(bid_note).strip():
        notes_array.append({"field": "Bid Note", "value": str(bid_note).strip()})

    return {
        "Customer Lane ID": f"LANE-{row_number}",
        "FO Code": "RXOCode",
        "Origin City": val("ORIGIN CITY", "ORIGIN_CITY", "ORIGIN"),
        "Origin State": val("ORIGIN STATE", "ORIGIN_STATE", "ORIGIN ST"),
        "Origin Zip": val("ORIGIN ZIP", "ORIGIN_ZIP", "ORIGIN ZIP CODE", "ORIGIN POSTAL", "ORIGIN_ZIP3"),
        "Origin Country": "USA" if str(origin_country).strip().upper() in {"US", "USA", "UNITED STATES"} else origin_country,
        "Destination City": val("DESTINATION CITY", "DESTINATION_CITY", "DESTINATION", "DEST CITY"),
        "Destination State": val("DESTINATION STATE", "DESTINATION_STATE", "DEST STATE", "DEST ST"),
        "Destination Zip": val("DESTINATION ZIP", "DESTINATION_ZIP", "DEST ZIP", "DEST ZIP CODE", "DEST POSTAL", "DEST_ZIP3"),
        "Destination Country": "USA" if str(dest_country).strip().upper() in {"US", "USA", "UNITED STATES"} else dest_country,
        "Origin Note": str(origin_note).strip() if origin_note else "",
        "Destination Note": str(dest_note).strip() if dest_note else "",
        "Bid Note": str(bid_note).strip() if bid_note else "",
        "Annual Volume": val("SHIPMENT COUNT", "LOADS PER YEAR", "ANNUAL VOLUME", "VOLUME"),
        "Equipment Category": "V",
        "Rate Type": "5T3",
        "Fuel Surcharge": "0.78",
        "FSC Type": "PerMileAmount",
        "Notes JSON": json.dumps(notes_array),
    }

def transform(context: dict) -> dict:
    workbook_path = context["input_workbook_path"]
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        records = []
        relevant_sheets = []
        ignored_sheets = []

        for sheet_name in workbook.sheetnames:
            upper_name = sheet_name.strip().upper()
            if any(token in upper_name for token in ["REQUIREMENT", "FSC", "LOCATION"]):
                ignored_sheets.append(sheet_name)
                continue

            ws = workbook[sheet_name]
            if ws.max_row < 2 or ws.max_column < 2:
                ignored_sheets.append(sheet_name)
                continue

            header_row_idx = None
            index_map = {}

            scan_rows = min(20, ws.max_row)
            for candidate in range(1, scan_rows + 1):
                values = [ws.cell(row=candidate, column=col).value for col in range(1, ws.max_column + 1)]
                norm = [_norm_header(v) for v in values]
                if (
                    "ORIGIN CITY" in norm
                    or "ORIGIN_CITY" in norm
                    or "ORIGIN_STATE" in norm
                ) and (
                    "DESTINATION CITY" in norm
                    or "DESTINATION_CITY" in norm
                    or "DEST CITY" in norm
                    or "DEST_ZIP3" in norm
                    or "DESTINATION_STATE" in norm
                ):
                    header_row_idx = candidate
                    index_map = {name: idx for idx, name in enumerate(norm)}
                    break

            if header_row_idx is None:
                ignored_sheets.append(sheet_name)
                continue

            relevant_sheets.append(sheet_name)
            for row_number in range(header_row_idx + 1, ws.max_row + 1):
                row = [ws.cell(row=row_number, column=col).value for col in range(1, ws.max_column + 1)]
                record = _row_to_record(row, index_map, row_number)
                if (
                    not record["Origin City"]
                    and not record["Destination City"]
                    and not record["Origin State"]
                    and not record["Destination State"]
                    and not record["Origin Zip"]
                    and not record["Destination Zip"]
                ):
                    continue
                records.append(record)

        notes = []
        for idx, rec in enumerate(records):
            for note_field in ("Origin Note", "Destination Note", "Bid Note"):
                note_value = rec.get(note_field, "")
                if note_value and str(note_value).strip():
                    notes.append({
                        "category": "carrier" if note_field == "Bid Note" else "special_handling",
                        "source_sheet": relevant_sheets[0] if relevant_sheets else "",
                        "source_column": note_field,
                        "note": str(note_value).strip(),
                        "severity": "info",
                        "lane_id": rec.get("Customer Lane ID", f"LANE-{idx+1}"),
                    })

        return {
            "records": records,
            "notes": notes,
            "metadata": {
                "relevant_sheets": relevant_sheets,
                "ignored_sheets": ignored_sheets,
                "record_count": len(records),
                "note_count": len(notes),
            },
        }
    finally:
        workbook.close()
""",
            "tests": ["assert 'FO Code' in output_columns"],
            "notes_json": [
                {
                    "category": "mapping",
                    "source_sheet": "",
                    "source_column": "",
                    "note": "Customer Lane ID is derived from row number (LANE-{row_number}).",
                    "severity": "info",
                },
                {
                    "category": "assumption",
                    "source_sheet": "",
                    "source_column": "",
                    "note": "Equipment Category defaulted to 'V' (Van) for all lanes.",
                    "severity": "info",
                },
                {
                    "category": "assumption",
                    "source_sheet": "",
                    "source_column": "",
                    "note": "Origin/Destination Country defaults to 'USA' when not provided.",
                    "severity": "warning",
                },
                {
                    "category": "data_quality",
                    "source_sheet": "",
                    "source_column": "ORIGIN NOTE|DESTINATION NOTE|BID NOTE",
                    "note": "Note/comment columns are preserved as-is from source data.",
                    "severity": "info",
                },
            ],
        }

    def _live_response(self, system_prompt: str, user_prompt: str) -> dict[str, Any]:
        if not self.endpoint:
            raise ValueError("FOUNDRY_PROJECT_ENDPOINT is required for live mode.")

        if self.agent_version:
            return self._live_response_via_agent_reference(system_prompt, user_prompt)

        prefers_assistants = bool(self.assistant_id.strip()) or self.agent_name.startswith("asst_")
        if prefers_assistants:
            return self._live_response_via_assistants(system_prompt, user_prompt)

        if "/api/projects/" in self.endpoint:
            try:
                return self._live_response_via_agent_reference(system_prompt, user_prompt)
            except Exception:
                return self._live_response_via_assistants(system_prompt, user_prompt)

        payload = {
            "agent_name": self.agent_name,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "response_format": {"type": "json_object"},
        }

        last_error: Exception | None = None
        for endpoint_url in self._build_live_url_candidates():
            for attempt in range(1, 3):
                try:
                    request = urllib.request.Request(
                        url=endpoint_url,
                        data=json.dumps(payload).encode("utf-8"),
                        headers=self._build_live_headers(),
                        method="POST",
                    )

                    with urllib.request.urlopen(request, timeout=60) as response:
                        raw = response.read().decode("utf-8")

                    decoded = json.loads(raw)
                    return self._extract_json_payload(decoded)
                except urllib.error.HTTPError as error:
                    body = ""
                    try:
                        body = error.read().decode("utf-8")
                    except Exception:
                        body = ""
                    last_error = RuntimeError(
                        f"HTTP {error.code} calling {endpoint_url}: {body or str(error)}"
                    )
                    if error.code in {400, 401, 403, 404}:
                        break
                    if attempt < 2:
                        time.sleep(1.5 * attempt)
                except Exception as error:
                    last_error = error
                    if attempt < 2:
                        time.sleep(1.5 * attempt)

        raise RuntimeError(f"Foundry planner call failed after retries: {last_error}")

    def _live_response_via_agent_reference(self, system_prompt: str, user_prompt: str) -> dict[str, Any]:
        if not self.agent_name.strip():
            raise ValueError("FOUNDRY_AGENT_NAME is required for agent_reference mode.")

        try:
            from azure.ai.projects import AIProjectClient
        except Exception as error:
            raise RuntimeError(
                "azure-ai-projects is required for New Foundry agent_reference mode. "
                "Install azure-ai-projects>=2.0.0."
            ) from error

        credential = (
            AzureKeyCredential(self.api_key)
            if self.api_key.strip()
            else DefaultAzureCredential(exclude_interactive_browser_credential=False)
        )

        project_client = AIProjectClient(endpoint=self.endpoint, credential=credential)
        openai_client = project_client.get_openai_client()

        agent_reference: dict[str, Any] = {
            "name": self.agent_name,
            "type": "agent_reference",
        }
        if self.agent_version:
            agent_reference["version"] = self.agent_version

        response = openai_client.responses.create(
            input=[
                {
                    "role": "user",
                    "content": (
                        "Follow these planner instructions exactly.\n\n"
                        + system_prompt
                        + "\n\nUser request context:\n"
                        + user_prompt
                    ),
                }
            ],
            extra_body={"agent_reference": agent_reference},
        )

        output_text = getattr(response, "output_text", None)
        if isinstance(output_text, str):
            parsed = self._try_parse_json_text(output_text)
            if parsed is not None:
                self.last_invocation_report = {
                    "mode": "live",
                    "path": "agent_reference",
                    "agent_name": self.agent_name,
                    "agent_version": self.agent_version,
                    "assistant_id": self.assistant_id,
                    "endpoint": self.endpoint,
                }
                return parsed

        if hasattr(response, "model_dump"):
            parsed = self._extract_json_payload(response.model_dump())
            if isinstance(parsed, dict):
                self.last_invocation_report = {
                    "mode": "live",
                    "path": "agent_reference",
                    "agent_name": self.agent_name,
                    "agent_version": self.agent_version,
                    "assistant_id": self.assistant_id,
                    "endpoint": self.endpoint,
                }
                return parsed

        raise RuntimeError("Unable to parse JSON payload from agent_reference response.")

    def _live_response_via_assistants(self, system_prompt: str, user_prompt: str) -> dict[str, Any]:
        api_version = os.getenv("FOUNDRY_API_VERSION", "2025-05-15-preview")
        base_url = self.endpoint.rstrip("/")
        assistant_id = self.assistant_id.strip()

        if not assistant_id:
            assistant_id = self._get_or_create_assistant_id(base_url, api_version, system_prompt)
        self.last_invocation_report = {
            "mode": "live",
            "path": "assistants",
            "agent_name": self.agent_name,
            "agent_version": self.agent_version,
            "assistant_id": assistant_id,
            "endpoint": self.endpoint,
        }

        run_payload = {
            "assistant_id": assistant_id,
            "thread": {
                "messages": [
                    {"role": "user", "content": user_prompt},
                ]
            },
        }
        run_result = self._post_json(
            f"{base_url}/threads/runs?api-version={api_version}",
            run_payload,
        )

        thread_id = run_result.get("thread_id")
        run_id = run_result.get("id")
        if not isinstance(thread_id, str) or not isinstance(run_id, str):
            raise RuntimeError("Foundry assistant run did not return thread_id/run_id.")

        for _ in range(60):
            status_payload = self._get_json(
                f"{base_url}/threads/{thread_id}/runs/{run_id}?api-version={api_version}"
            )
            status = status_payload.get("status")
            if status in {"completed", "failed", "cancelled", "expired"}:
                if status != "completed":
                    raise RuntimeError(f"Foundry assistant run ended with status={status}")
                break
            time.sleep(1)

        messages_payload = self._get_json(
            f"{base_url}/threads/{thread_id}/messages?api-version={api_version}"
        )
        parsed = self._extract_assistant_json_from_messages(messages_payload)
        if parsed is None:
            raise RuntimeError("Unable to parse JSON from assistant response message.")
        return parsed

    def _extract_assistant_json_from_messages(self, messages_payload: dict[str, Any]) -> dict[str, Any] | None:
        data = messages_payload.get("data")
        if not isinstance(data, list):
            return None

        for message in data:
            if not isinstance(message, dict):
                continue
            if message.get("role") != "assistant":
                continue

            content = message.get("content")
            if not isinstance(content, list):
                continue

            for chunk in content:
                if not isinstance(chunk, dict):
                    continue
                if chunk.get("type") != "text":
                    continue
                text_block = chunk.get("text")
                if not isinstance(text_block, dict):
                    continue
                text_value = text_block.get("value")
                if isinstance(text_value, str):
                    parsed = self._try_parse_json_text(text_value)
                    if parsed is not None:
                        return parsed

        return None

    def _try_parse_json_text(self, text_value: str) -> dict[str, Any] | None:
        candidates = [text_value]

        stripped = text_value.strip()
        if stripped.startswith("```"):
            lines = stripped.splitlines()
            if len(lines) >= 3:
                candidates.append("\n".join(lines[1:-1]))

        first_brace = text_value.find("{")
        last_brace = text_value.rfind("}")
        if first_brace != -1 and last_brace != -1 and last_brace > first_brace:
            candidates.append(text_value[first_brace:last_brace + 1])

        for candidate in candidates:
            try:
                parsed = json.loads(candidate)
                if isinstance(parsed, dict):
                    return parsed
            except Exception:
                continue

        return None

    def _get_or_create_assistant_id(self, base_url: str, api_version: str, system_prompt: str) -> str:
        if self.agent_name.startswith("asst_"):
            return self.agent_name

        assistants = self._get_json(f"{base_url}/assistants?api-version={api_version}")
        for item in assistants.get("data", []):
            if isinstance(item, dict) and item.get("name") == self.agent_name and isinstance(item.get("id"), str):
                return item["id"]

        model_name = os.getenv("FOUNDRY_MODEL", "gpt-4.1")
        created = self._post_json(
            f"{base_url}/assistants?api-version={api_version}",
            {
                "name": self.agent_name or "RXO-Document-Normalizer-Assistant",
                "model": model_name,
                "instructions": system_prompt,
            },
        )
        assistant_id = created.get("id")
        if not isinstance(assistant_id, str) or not assistant_id:
            raise RuntimeError("Failed to create Foundry assistant for live mode.")
        return assistant_id

    def _get_json(self, url: str) -> dict[str, Any]:
        request = urllib.request.Request(
            url=url,
            headers=self._build_live_headers(),
            method="GET",
        )
        with urllib.request.urlopen(request, timeout=60) as response:
            return json.loads(response.read().decode("utf-8"))

    def _post_json(self, url: str, payload: dict[str, Any]) -> dict[str, Any]:
        request = urllib.request.Request(
            url=url,
            data=json.dumps(payload).encode("utf-8"),
            headers=self._build_live_headers(),
            method="POST",
        )
        with urllib.request.urlopen(request, timeout=60) as response:
            return json.loads(response.read().decode("utf-8"))

    def _build_live_url_candidates(self) -> list[str]:
        base_url = self.endpoint.strip()
        if not base_url:
            return []

        candidates: list[str] = [base_url]
        normalized = base_url.rstrip("/")
        api_versions_csv = os.getenv(
            "FOUNDRY_API_VERSIONS",
            "2025-05-15-preview,2025-04-01-preview,2024-10-21,2024-05-01-preview",
        )
        api_versions = [item.strip() for item in api_versions_csv.split(",") if item.strip()]

        if "/api/projects/" in normalized:
            for api_version in api_versions:
                candidates.append(f"{normalized}/chat/completions?api-version={api_version}")
                if self.agent_name:
                    candidates.append(
                        f"{normalized}/agents/{self.agent_name}/chat/completions?api-version={api_version}"
                    )

        unique_candidates: list[str] = []
        for candidate in candidates:
            if candidate not in unique_candidates:
                unique_candidates.append(candidate)
        return unique_candidates

    def _build_live_headers(self) -> dict[str, str]:
        headers = {
            "Content-Type": "application/json",
        }

        if self.api_key:
            headers["api-key"] = self.api_key
            return headers

        token = self._get_aad_bearer_token()
        headers["Authorization"] = f"Bearer {token}"
        return headers

    def _get_aad_bearer_token(self) -> str:
        token_scope = os.getenv(
            "FOUNDRY_TOKEN_SCOPE",
            "https://ai.azure.com/.default",
        )
        try:
            from azure.identity import DefaultAzureCredential
        except Exception as error:
            raise RuntimeError(
                "azure-identity is required for live mode when FOUNDRY_API_KEY is not set. "
                "Install azure-identity or provide FOUNDRY_API_KEY."
            ) from error

        credential = DefaultAzureCredential(exclude_interactive_browser_credential=False)
        token = credential.get_token(token_scope)
        return token.token

    def _extract_json_payload(self, decoded: Any) -> dict[str, Any]:
        if isinstance(decoded, dict):
            if isinstance(decoded.get("content"), str):
                return json.loads(decoded["content"])
            if isinstance(decoded.get("output_text"), str):
                return json.loads(decoded["output_text"])

            output = decoded.get("output")
            if isinstance(output, list):
                for item in output:
                    if not isinstance(item, dict):
                        continue
                    content_items = item.get("content")
                    if not isinstance(content_items, list):
                        continue
                    for content_item in content_items:
                        if not isinstance(content_item, dict):
                            continue
                        text_value = content_item.get("text")
                        if isinstance(text_value, str):
                            parsed = self._try_parse_json_text(text_value)
                            if parsed is not None:
                                return parsed

            choices = decoded.get("choices")
            if isinstance(choices, list) and choices:
                first_choice = choices[0]
                if isinstance(first_choice, dict):
                    message = first_choice.get("message")
                    if isinstance(message, dict):
                        content = message.get("content")
                        if isinstance(content, str):
                            return json.loads(content)
                        if isinstance(content, list):
                            for item in content:
                                if isinstance(item, dict) and item.get("type") in {"output_text", "text"}:
                                    text_value = item.get("text")
                                    if isinstance(text_value, str):
                                        return json.loads(text_value)

            if "relevant_sheets" in decoded and "python_script" in decoded:
                return decoded

        raise ValueError("Unable to extract planner JSON payload from live response.")
