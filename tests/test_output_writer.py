import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.function_app.services.output_writer import (
    get_canonical_columns,
    normalize_records_to_canonical,
    write_canonical_csv,
    write_canonical_xlsx,
)
from src.function_app.services.template_loader import load_canonical_schema


class TestOutputWriter(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.schema = load_canonical_schema(
            "src/function_app/templates/canonical_schema.freight_bid_v1.json"
        )

    def test_get_canonical_columns_preserves_template_order(self) -> None:
        columns = get_canonical_columns(self.schema)
        self.assertEqual(columns[0], "Customer Lane ID")
        self.assertEqual(columns[1], "FO Code")
        self.assertEqual(columns[-1], "Strategic Quantile")
        self.assertEqual(len(columns), 40)

    def test_normalize_records_fills_missing_columns(self) -> None:
        input_records = [
            {
                "Customer Lane ID": "L1",
                "FO Code": "RXOCode",
                "Origin City": "Dallas",
            }
        ]
        normalized = normalize_records_to_canonical(input_records, self.schema)

        self.assertEqual(len(normalized), 1)
        self.assertEqual(normalized[0]["Customer Lane ID"], "L1")
        self.assertEqual(normalized[0]["FO Code"], "RXOCode")
        self.assertEqual(normalized[0]["Origin City"], "Dallas")
        self.assertEqual(normalized[0]["Destination City"], "")

    def test_write_canonical_csv_writes_header_and_rows(self) -> None:
        records = [
            {
                "Customer Lane ID": "L1",
                "FO Code": "RXOCode",
                "Origin City": "Dallas",
                "Destination City": "Austin",
            }
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "out.csv"
            write_canonical_csv(records, self.schema, str(csv_path))

            self.assertTrue(csv_path.exists())
            with csv_path.open("r", encoding="utf-8", newline="") as file_handle:
                reader = csv.reader(file_handle)
                rows = list(reader)

        self.assertEqual(rows[0][0], "Customer Lane ID")
        self.assertEqual(rows[0][1], "FO Code")
        self.assertEqual(rows[1][0], "L1")
        self.assertEqual(rows[1][2], "Dallas")

    def test_write_canonical_xlsx_writes_header_and_rows(self) -> None:
        records = [
            {
                "Customer Lane ID": "L2",
                "FO Code": "RXOCode",
                "Origin City": "Houston",
                "Destination City": "El Paso",
            }
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            xlsx_path = Path(temp_dir) / "out.xlsx"
            write_canonical_xlsx(records, self.schema, str(xlsx_path))

            self.assertTrue(xlsx_path.exists())
            workbook = load_workbook(xlsx_path)
            worksheet = workbook["CanonicalOutput"]

            self.assertEqual(worksheet.cell(row=1, column=1).value, "Customer Lane ID")
            self.assertEqual(worksheet.cell(row=1, column=2).value, "FO Code")
            self.assertEqual(worksheet.cell(row=2, column=1).value, "L2")
            self.assertEqual(worksheet.cell(row=2, column=3).value, "Houston")

    def test_normalize_records_applies_country_and_bool_normalization(self) -> None:
        input_records = [
            {
                "Customer Lane ID": "L3",
                "Origin Country": "us",
                "Destination Country": "Canada",
                "Hazmat": "Y",
            }
        ]

        normalized = normalize_records_to_canonical(input_records, self.schema)

        self.assertEqual(normalized[0]["Origin Country"], "USA")
        self.assertEqual(normalized[0]["Destination Country"], "CAN")
        self.assertEqual(normalized[0]["Hazmat"], True)


if __name__ == "__main__":
    unittest.main()
